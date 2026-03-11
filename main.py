#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Jewelry Store Product Management Application - Windows Compatible Version
Enhanced with better UX, performance improvements, error handling, and clear text display
Compatible with all Windows systems without custom ttk styles

FIXED VERSION - إصدار محدث:
=================================
تم إصلاح مشكلة ترتيب الأعمدة عند الحفظ ومشكلة فقدان الأصفار في الباركود:

1. إضافة دالة validate_column_mapping() للتحقق من صحة ترتيب الأعمدة
2. تحسين دالة export_to_excel() للحفاظ على ترتيب الأعمدة الأصلي
3. إضافة فحوصات إضافية لضمان سلامة البيانات عند الحفظ
4. التأكد من عدم تغيير ترتيب الأعمدة أو البيانات غير المعدلة
5. إصلاح مشكلة فقدان الأصفار في بداية الباركود

التحديثات الرئيسية:
- تحسين آلية column mapping
- إضافة فحوصات للتأكد من صحة البيانات
- الحفاظ على ترتيب الأعمدة الأصلي بدقة
- منع فقدان أو تغيير البيانات غير المعدلة
- الحفاظ على الأصفار في بداية الباركود (مثل 001234)
- معالجة خاصة للباركود عند القراءة والكتابة
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image, ImageTk
import os
import glob
from datetime import datetime
import json
import threading
import queue
from pathlib import Path
import platform
from typing import List, Dict, Optional, Tuple
import logging

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ImageCache:
    """Cache for thumbnail images to improve performance"""

    def __init__(self, max_size: int = 100):
        self.cache = {}
        self.max_size = max_size

    def get(self, path: str, size: Tuple[int, int]) -> Optional[ImageTk.PhotoImage]:
        key = f"{path}_{size[0]}x{size[1]}"
        return self.cache.get(key)

    def set(self, path: str, size: Tuple[int, int], image: ImageTk.PhotoImage):
        key = f"{path}_{size[0]}x{size[1]}"
        if len(self.cache) >= self.max_size:
            # Remove oldest item (simple FIFO)
            self.cache.pop(next(iter(self.cache)))
        self.cache[key] = image


class ProgressDialog:
    """Progress dialog for long operations"""

    def __init__(self, parent, title="جاري المعالجة", message="يرجى الانتظار..."):
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("300x120")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        # Center the dialog
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - 150
        y = (self.dialog.winfo_screenheight() // 2) - 60
        self.dialog.geometry(f"+{x}+{y}")

        # Message label
        self.message_label = ttk.Label(self.dialog, text=message, font=('Arial', 10))
        self.message_label.pack(pady=20)

        # Progress bar
        self.progress = ttk.Progressbar(self.dialog, mode='indeterminate', length=250)
        self.progress.pack(pady=10)
        self.progress.start(10)

        # Prevent closing
        self.dialog.protocol("WM_DELETE_WINDOW", lambda: None)

    def update_message(self, message: str):
        self.message_label.config(text=message)
        self.dialog.update()

    def close(self):
        self.progress.stop()
        self.dialog.destroy()


class JewelryStoreManager:
    def __init__(self, root):
        self.root = root
        self.root.title("إدارة منتجات متجر المجوهرات - Jewelry Store Manager Pro")
        self.root.geometry("1400x800")

        # Beautiful purple theme colors
        self.purple_color = '#8B4A9C'  # Main purple from the image
        self.light_purple = '#B47CC7'  # Lighter shade
        self.gold_accent = '#D4AF37'  # Gold accent color
        self.pink_highlight = '#FFB6C1'  # Light pink for selected product highlighting
        self.root.configure(bg=self.purple_color)

        # Set minimum window size
        self.root.minsize(1200, 600)

        # Configure font for better Arabic text display
        self.setup_fonts()

        # Initialize image cache
        self.image_cache = ImageCache()

        # Track if data has been modified
        self.data_modified = False

        # Undo/Redo stacks
        self.undo_stack = []
        self.redo_stack = []
        self.max_undo_size = 20

        # Arabic column names from the Excel template
        self.columns = [
            'أسم المنتج',  # Product Name
            'تصنيف المنتج',  # Product Category
            'صورة المنتج',  # Product Image
            'وصف صورة المنتج',  # Product Image Description
            'سعر المنتج',  # Product Price
            'الكمية المتوفرة',  # Available Quantity
            'رمز المنتج sku',  # Product SKU
            'الوزن',  # Weight
            'وحدة الوزن',  # Weight Unit
            'الباركود',  # Barcode
            'المصنعية للجرام',  # Craftsmanship per Gram
            'المصنعية للقطعة',  # Craftsmanship per Piece
            'الكمية المباعة',  # Sold Quantity
            'العيار',  # Karat/Grade
            'وزن الفصوص',  # Gemstone Weight
            'المقاس',  # Size
            'المورد',  # Supplier
            'رقم الفاتورة',  # Invoice Number
            'تاريخ الفاتورة'  # Invoice Date
        ]

        # Product categories for dropdown
        self.product_categories = [
            'خواتم',
            'بناجر',
            'إنسيالات',
            'كفوف',
            'عقد',
            'حلق',
            'أطقم أساور وخواتم',
            'أنصاف أطقم',
            'أطقم',
            'خلخال',
            'أطفال'
        ]

        # Initialize data structures
        self.products_data = []
        self.current_folder = ""
        self.current_excel_file = None  # Track the current Excel file being used
        self.original_column_mapping = {}  # Map program columns to original file column positions
        self.original_headers = []  # Store original file headers
        self.image_files = []
        self.selected_images = {}
        self.current_product_index = -1
        self.image_widgets = []
        self.selected_image_for_removal = -1

        # Column visibility control
        self.visible_columns = {col: True for col in self.columns}  # All columns visible by default
        self.load_column_preferences()  # Load saved preferences

        # Default values persistence
        self.default_values = {
            'المورد': '',
            'رقم الفاتورة': '',
            'تاريخ الفاتورة': '',
            'العيار': '',
            'المصنعية للجرام': '',
            'المصنعية للقطعة': ''
        }

        # Image selection tracking for all products
        self.all_selected_images = {}

        # Search variables
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.on_search_changed)
        self.is_searching = False  # Track if we're in search mode

        # Advanced search functionality
        self.search_results = []
        self.current_search_index = -1
        self.last_search_term = ""

        # Store original data for search filtering
        self.original_products_data = []
        self.filtered_indices = []  # Track which original indices are currently shown

        # Create the main interface
        self.create_widgets()
        self.setup_layout()
        self.load_default_values()
        self.setup_keyboard_shortcuts()

        # Bind window close event
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        # Auto-save timer (every 5 minutes)
        self.auto_save_interval = 300000  # 5 minutes in milliseconds
        self.schedule_auto_save()

    def setup_fonts(self):
        """Setup fonts for better Arabic text display"""
        try:
            # Configure default font for better Arabic support
            default_font = ('Segoe UI', 9) if platform.system() == "Windows" else ('Arial', 9)
            self.root.option_add('*Font', default_font)

            # Configure specific fonts for different elements
            self.header_font = ('Segoe UI', 10, 'bold') if platform.system() == "Windows" else ('Arial', 10, 'bold')
            self.table_font = ('Segoe UI', 9) if platform.system() == "Windows" else ('Arial', 9)
            self.button_font = ('Segoe UI', 9) if platform.system() == "Windows" else ('Arial', 9)

        except Exception as e:
            logger.error(f"Font setup error: {e}")
            # Fallback fonts
            self.header_font = ('Arial', 10, 'bold')
            self.table_font = ('Arial', 9)
            self.button_font = ('Arial', 9)

    def setup_keyboard_shortcuts(self):
        """Setup keyboard shortcuts"""
        # Keyboard shortcuts
        self.root.bind('<Control-s>', lambda e: self.save_data())
        self.root.bind('<Control-z>', lambda e: self.undo())
        self.root.bind('<Control-y>', lambda e: self.redo())
        self.root.bind('<Control-f>', lambda e: self.show_advanced_search())  # Advanced search
        self.root.bind('<F3>', lambda e: self.find_next())  # Find next
        self.root.bind('<Shift-F3>', lambda e: self.find_previous())  # Find previous
        self.root.bind('<Control-n>', lambda e: self.add_new_product())
        self.root.bind('<Delete>', lambda e: self.delete_selected_image())
        self.root.bind('<Control-a>', lambda e: self.select_all_images())
        self.root.bind('<Escape>', lambda e: self.clear_selection())

    def schedule_auto_save(self):
        """Schedule automatic saving"""
        if self.data_modified and self.current_folder:
            self.auto_save()
        self.root.after(self.auto_save_interval, self.schedule_auto_save)

    def auto_save(self):
        """Perform auto-save to a backup file"""
        try:
            if not self.current_folder:
                return

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_dir = os.path.join(self.current_folder, "backups")
            os.makedirs(backup_dir, exist_ok=True)

            filename = f"Products_AutoSave_{timestamp}.xlsx"
            filepath = os.path.join(backup_dir, filename)

            # Save quietly without showing message
            self._save_to_file(filepath, show_message=False)

            # Clean old backups (keep only last 5)
            self._clean_old_backups(backup_dir, keep_count=5)

        except Exception as e:
            logger.error(f"Auto-save failed: {e}")

    def _clean_old_backups(self, backup_dir: str, keep_count: int = 5):
        """Clean old backup files"""
        try:
            backup_files = sorted(
                glob.glob(os.path.join(backup_dir, "Products_AutoSave_*.xlsx")),
                key=os.path.getmtime,
                reverse=True
            )

            for old_file in backup_files[keep_count:]:
                try:
                    os.remove(old_file)
                except:
                    pass
        except Exception as e:
            logger.error(f"Failed to clean old backups: {e}")

    def save_state_for_undo(self):
        """Save current state for undo functionality"""
        state = {
            'products_data': [row[:] for row in self.products_data],
            'all_selected_images': {k: v[:] for k, v in self.all_selected_images.items()}
        }

        self.undo_stack.append(state)
        if len(self.undo_stack) > self.max_undo_size:
            self.undo_stack.pop(0)

        # Clear redo stack when new action is performed
        self.redo_stack.clear()

    def undo(self):
        """Undo last action"""
        if not self.undo_stack:
            return

        # Save current state to redo stack
        current_state = {
            'products_data': [row[:] for row in self.products_data],
            'all_selected_images': {k: v[:] for k, v in self.all_selected_images.items()}
        }
        self.redo_stack.append(current_state)

        # Restore previous state
        state = self.undo_stack.pop()
        self.restore_state(state)

    def redo(self):
        """Redo last undone action"""
        if not self.redo_stack:
            return

        # Save current state to undo stack
        self.save_state_for_undo()

        # Restore redo state
        state = self.redo_stack.pop()
        self.restore_state(state)

    def restore_state(self, state):
        """Restore a saved state"""
        self.products_data = [row[:] for row in state['products_data']]
        self.all_selected_images = {k: v[:] for k, v in state['all_selected_images'].items()}

        # Refresh UI
        self.refresh_product_table()
        self.display_selected_images()
        self.update_all_image_borders()

    def on_closing(self):
        """Handle window closing event"""
        if self.data_modified:
            result = messagebox.askyesnocancel(
                "حفظ التغييرات",
                "هل تريد حفظ التغييرات قبل إغلاق البرنامج؟\n\nنعم: حفظ وإغلاق\nلا: إغلاق بدون حفظ\nإلغاء: العودة للبرنامج"
            )

            if result is True:  # Yes - Save and close
                if self.save_data():
                    self.root.destroy()
            elif result is False:  # No - Close without saving
                self.root.destroy()
            # Cancel - Do nothing
        else:
            self.root.destroy()

    def mark_data_modified(self):
        """Mark data as modified"""
        self.data_modified = True
        if not self.root.title().endswith(" *"):
            self.root.title(self.root.title() + " *")

    def mark_data_saved(self):
        """Mark data as saved"""
        self.data_modified = False
        title = self.root.title()
        if title.endswith(" *"):
            self.root.title(title[:-2])

    def create_widgets(self):
        """Create all the GUI widgets"""
        # Main frame with purple theme
        self.main_frame = tk.Frame(self.root, bg=self.purple_color)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Top toolbar
        self.create_toolbar()

        # Three-section layout
        self.create_three_sections()

        # Status bar
        self.create_status_bar()

    def create_toolbar(self):
        """Create the top toolbar with improved layout"""
        toolbar = tk.Frame(self.main_frame, bg=self.purple_color)
        toolbar.pack(fill=tk.X, pady=(0, 10))

        # Left side - Folder operations
        left_frame = tk.Frame(toolbar, bg=self.purple_color)
        left_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        folder_btn = tk.Button(left_frame, text="📁 اختيار مجلد الصور",
                               command=self.select_folder,
                               bg=self.light_purple, fg='white',
                               font=self.button_font, relief='raised', bd=2)
        folder_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.folder_label = tk.Label(left_frame, text="لم يتم اختيار مجلد",
                                     foreground="white", bg=self.purple_color, font=self.table_font)
        self.folder_label.pack(side=tk.LEFT, padx=(0, 20))

        # Search box
        search_frame = tk.Frame(left_frame, bg=self.purple_color)
        search_frame.pack(side=tk.LEFT, padx=(20, 0))

        tk.Label(search_frame, text="🔍 بحث:", font=self.table_font,
                 bg=self.purple_color, fg='white').pack(side=tk.LEFT, padx=(0, 5))
        self.search_entry = tk.Entry(search_frame, textvariable=self.search_var, width=20, font=self.table_font)
        self.search_entry.pack(side=tk.LEFT)

        # Clear search button
        clear_search_btn = tk.Button(search_frame, text="✖",
                                     command=self.clear_search_filter,
                                     bg='#FF6B6B', fg='white',
                                     font=self.button_font, width=2, height=1)
        clear_search_btn.pack(side=tk.LEFT, padx=(2, 0))

        # Tooltip for clear button
        self._create_tooltip_for_widget(clear_search_btn, "مسح البحث وإظهار جميع المنتجات")

        # Right side - Action buttons
        right_frame = tk.Frame(toolbar, bg=self.purple_color)
        right_frame.pack(side=tk.RIGHT)

        # Create buttons with purple theme
        buttons = [
            ("↶ تراجع", self.undo),
            ("↷ إعادة", self.redo),
            ("⚙️ إعدادات الأعمدة", self.show_column_settings),
            ("➕ منتج جديد", self.add_new_product),
            ("💾 حفظ", self.save_data),
            ("📤 تصدير", self.export_data)
        ]

        for i, (text, command) in enumerate(buttons):
            if i == 3:  # Add separator before "منتج جديد"
                separator = tk.Frame(right_frame, width=2, bg='white')
                separator.pack(side=tk.LEFT, fill=tk.Y, padx=5)

            btn = tk.Button(right_frame, text=text, command=command,
                            bg=self.light_purple, fg='white',
                            font=self.button_font, relief='raised', bd=2)
            btn.pack(side=tk.LEFT, padx=2)

    def create_status_bar(self):
        """Create status bar at the bottom"""
        self.status_bar = tk.Frame(self.root, bg=self.purple_color)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=(0, 5))

        # Product count
        self.product_count_label = tk.Label(self.status_bar, text="المنتجات: 0",
                                            font=self.table_font, bg=self.purple_color, fg='white')
        self.product_count_label.pack(side=tk.LEFT, padx=(0, 20))

        # Image count
        self.image_count_label = tk.Label(self.status_bar, text="الصور: 0",
                                          font=self.table_font, bg=self.purple_color, fg='white')
        self.image_count_label.pack(side=tk.LEFT, padx=(0, 20))

        # Selected images count
        self.selected_count_label = tk.Label(self.status_bar, text="المحددة: 0",
                                             font=self.table_font, bg=self.purple_color, fg='white')
        self.selected_count_label.pack(side=tk.LEFT)

        # Toast message area
        self.toast_label = tk.Label(self.status_bar, text="",
                                    font=self.table_font, bg=self.purple_color, fg=self.gold_accent)
        self.toast_label.pack(side=tk.RIGHT)

    def show_toast(self, title, message, duration=3000):
        """Show toast notification"""
        self.toast_label.config(text=f"{title}: {message}")
        self.root.after(duration, lambda: self.toast_label.config(text=""))

    def update_status_bar(self):
        """Update status bar information"""
        # Product count
        product_count = len(self.products_data)
        self.product_count_label.config(text=f"المنتجات: {product_count}")

        # Image count
        image_count = len(self.image_files)
        self.image_count_label.config(text=f"الصور: {image_count}")

        # Selected images count for current product
        selected_count = 0
        if self.current_product_index >= 0 and self.current_product_index in self.all_selected_images:
            selected_count = len(self.all_selected_images[self.current_product_index])
        self.selected_count_label.config(text=f"المحددة: {selected_count}")

    def copy_to_clipboard(self, text):
        """Copy text to clipboard"""
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.show_toast("تم النسخ", "تم نسخ النص للحافظة")

    def create_three_sections(self):
        """Create the three main sections of the interface with resizable panes"""
        # Main PanedWindow for resizable sections
        main_paned = tk.PanedWindow(self.main_frame, orient=tk.HORIZONTAL,
                                    sashrelief=tk.RAISED, sashwidth=6,
                                    bg=self.purple_color, sashpad=2)
        main_paned.pack(fill=tk.BOTH, expand=True)

        # Left section - Images
        left_section = self.create_left_section(main_paned)
        main_paned.add(left_section, minsize=250, width=350)  # Minimum 250px, default 350px

        # Create a sub-paned window for middle and right sections
        right_paned = tk.PanedWindow(main_paned, orient=tk.HORIZONTAL,
                                     sashrelief=tk.RAISED, sashwidth=6,
                                     bg=self.purple_color, sashpad=2)
        main_paned.add(right_paned, minsize=600)  # Minimum width for right paned window

        # Middle section - Selected images
        middle_section = self.create_middle_section(right_paned)
        right_paned.add(middle_section, minsize=250, width=300)  # Minimum 250px, default 300px

        # Right section - Product table
        right_section = self.create_right_section(right_paned)
        right_paned.add(right_section, minsize=400)  # Minimum 400px for table

        # Store references for later use
        self.main_paned = main_paned
        self.right_paned = right_paned

    def create_left_section(self, parent):
        """Create the left section for image display"""
        left_frame = tk.LabelFrame(parent, text="صور المجلد",
                                   bg=self.purple_color, fg='white',
                                   font=self.header_font, bd=2, relief='raised')

        # Scrollable canvas for images
        canvas = tk.Canvas(left_frame, width=280, bg='white')
        scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=canvas.yview)
        self.scrollable_frame = tk.Frame(canvas, bg='white')

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        scrollbar.pack(side="right", fill="y")

        self.images_canvas = canvas
        self._bind_mousewheel(canvas)

        return left_frame

    def create_middle_section(self, parent):
        """Create the middle section for selected images display"""
        middle_frame = tk.LabelFrame(parent, text="صور المنتج المحدد",
                                     bg=self.purple_color, fg='white',
                                     font=self.header_font, bd=2, relief='raised')

        # Product info label
        self.selected_info_label = tk.Label(middle_frame, text="لم يتم اختيار منتج",
                                            font=self.header_font, foreground=self.gold_accent,
                                            background=self.purple_color)
        self.selected_info_label.pack(pady=(5, 10))

        # Selected images display with scrollbar
        canvas = tk.Canvas(middle_frame, bg='white', height=400)
        scrollbar = ttk.Scrollbar(middle_frame, orient="vertical", command=canvas.yview)
        self.selected_images_frame = tk.Frame(canvas, bg='white')

        self.selected_images_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.selected_images_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True, padx=5)
        scrollbar.pack(side="right", fill="y")

        self.selected_canvas = canvas
        self._bind_mousewheel(canvas)

        # Buttons for image management
        buttons_frame = tk.Frame(middle_frame, bg=self.purple_color)
        buttons_frame.pack(fill=tk.X, pady=(10, 5), padx=5)

        # Remove button
        remove_btn = tk.Button(buttons_frame, text="❌ إزالة الصورة المحددة",
                               command=self.remove_selected_image,
                               bg=self.light_purple, fg='white', font=self.button_font)
        remove_btn.pack(fill=tk.X, pady=(0, 5))

        # Reorder button
        reorder_btn = tk.Button(buttons_frame, text="🔄 ترتيب الصور",
                                command=self.reorder_images,
                                bg=self.light_purple, fg='white', font=self.button_font)
        reorder_btn.pack(fill=tk.X, pady=(0, 5))

        # Instructions label
        instructions = tk.Label(buttons_frame,
                                text="انقر على صورة لتحديدها\nزر الماوس الأيمن لنسخ مسار الصورة",
                                font=('Arial', 8), foreground='white', bg=self.purple_color, justify=tk.CENTER)
        instructions.pack(pady=(5, 0))

        return middle_frame

    def create_right_section(self, parent):
        """Create the right section for product data table with enhanced text display"""
        right_frame = tk.LabelFrame(parent, text="بيانات المنتجات",
                                    bg=self.purple_color, fg='white',
                                    font=self.header_font, bd=2, relief='raised')

        # Create treeview for product data
        self.create_product_table(right_frame)

        return right_frame

    def create_product_table(self, parent):
        """Create enhanced product data table with clear text display"""
        # Frame for table and scrollbars
        table_frame = tk.Frame(parent, bg=self.purple_color)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Create treeview with enhanced styling for better text display
        style = ttk.Style()
        style.configure("Enhanced.Treeview",
                        rowheight=50,  # Doubled row height for perfect Arabic text visibility
                        font=self.table_font)
        style.configure("Enhanced.Treeview.Heading",
                        font=self.header_font,
                        background='#E1E1E1')
        style.map('Enhanced.Treeview',
                  background=[('selected', '#0078D7')],
                  foreground=[('selected', 'white')])

        # Create treeview with visible columns only
        visible_cols = [col for col in self.columns if self.visible_columns[col]]
        self.tree = ttk.Treeview(table_frame, columns=visible_cols, show='headings',
                                 height=15, style="Enhanced.Treeview")

        # Configure alternating row colors for better readability
        self.tree.tag_configure('oddrow', background='#F8F8F8')
        self.tree.tag_configure('evenrow', background='white')

        # Configure tags for highlighting
        self.tree.tag_configure('has_images', background='#FFB6C1')  # Light pink for rows with images
        self.tree.tag_configure('selected_product',
                                background=self.pink_highlight)  # Pink highlight for selected product

        # Configure columns with optimized widths for Arabic text
        column_widths = {
            'أسم المنتج': 250,  # Increased from 180
            'تصنيف المنتج': 180,  # Increased from 120
            'صورة المنتج': 200,  # Increased from 150
            'وصف صورة المنتج': 250,  # Increased from 180
            'سعر المنتج': 120,  # Increased from 100
            'الكمية المتوفرة': 120,  # Increased from 100
            'رمز المنتج sku': 150,  # Increased from 120
            'الوزن': 100,  # Increased from 80
            'وحدة الوزن': 100,  # Increased from 80
            'الباركود': 150,  # Increased from 120
            'المصنعية للجرام': 140,  # Increased from 110
            'المصنعية للقطعة': 140,  # Increased from 110
            'الكمية المباعة': 120,  # Increased from 100
            'العيار': 100,  # Increased from 80
            'وزن الفصوص': 120,  # Increased from 100
            'المقاس': 100,  # Increased from 80
            'المورد': 150,  # Increased from 120
            'رقم الفاتورة': 150,  # Increased from 120
            'تاريخ الفاتورة': 150  # Increased from 120
        }

        # Configure visible columns only with resizable option
        for col in visible_cols:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_table(c))
            width = column_widths.get(col, 100)
            # Enable column resizing but remove stretch to allow horizontal scrolling
            self.tree.column(col, width=width, minwidth=50, anchor='center', stretch=False)

        # Add double-click auto-resize functionality
        self.setup_column_auto_resize()

        # Scrollbars
        v_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Grid layout
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Bind events
        self.tree.bind('<<TreeviewSelect>>', self.on_product_select)
        # DISABLED: Direct editing removed for safety - use right-click menu instead
        # self.tree.bind('<Double-1>', self.edit_product)  # REMOVED: Direct editing disabled
        self.tree.bind('<Button-3>', self.show_context_menu)  # Right-click menu
        self.tree.bind('<KeyRelease>', self.on_tree_change)  # Track changes
        self.tree.bind('<FocusOut>', self.on_tree_change)  # Track focus changes

        # DISABLED: Direct editing removed for safety
        # self.setup_tree_editing()  # REMOVED: Direct editing disabled

        # Sort state
        self.sort_reverse = False

    def setup_column_auto_resize(self):
        """Setup double-click auto-resize for column headers"""

        def on_header_double_click(event):
            # Get the column that was double-clicked
            region = self.tree.identify_region(event.x, event.y)
            if region == "heading":
                column = self.tree.identify_column(event.x)
                if column:
                    # Convert column identifier to column name
                    col_index = int(column[1:]) - 1  # Remove '#' and convert to 0-based
                    visible_cols = [col for col in self.columns if self.visible_columns[col]]

                    if 0 <= col_index < len(visible_cols):
                        col_name = visible_cols[col_index]
                        self.auto_resize_column(col_name)

        # Bind double-click event to tree
        self.tree.bind('<Double-Button-1>', on_header_double_click)

    def auto_resize_column(self, column_name):
        """Auto-resize column to fit content"""
        try:
            # Get all visible columns
            visible_cols = [col for col in self.columns if self.visible_columns[col]]

            if column_name not in visible_cols:
                return

            col_index = visible_cols.index(column_name)

            # Calculate maximum width needed
            max_width = 100  # Minimum width

            # Check header width
            header_width = len(column_name) * 12  # Approximate character width
            max_width = max(max_width, header_width)

            # Check all data in this column
            for item in self.tree.get_children():
                values = self.tree.item(item)['values']
                if col_index < len(values):
                    cell_value = str(values[col_index])
                    # Calculate approximate width (Arabic text needs more space)
                    cell_width = len(cell_value) * 15  # Increased for Arabic text
                    max_width = max(max_width, cell_width)

            # Add some padding
            max_width += 20

            # Set maximum reasonable width
            max_width = min(max_width, 400)

            # Apply the new width
            self.tree.column(f"#{col_index + 1}", width=max_width)

            # Show feedback
            self.show_toast("تم التوسيع", f"تم توسيع عمود '{column_name}' تلقائياً")

        except Exception as e:
            print(f"Error in auto_resize_column: {e}")

    def setup_tree_editing(self):
        """DISABLED: Direct editing removed for safety - use right-click menu instead"""
        # REMOVED: All direct editing functionality disabled
        # Users must use right-click menu -> Edit to modify products
        pass

    def start_edit(self, item, col_index):
        """DISABLED: Direct editing removed for safety - use right-click menu instead"""
        # REMOVED: Direct editing functionality disabled
        # Users must use right-click menu -> Edit to modify products
        return

    def finish_edit(self):
        """Finish editing and save changes"""
        if not self.edit_entry or not self.edit_item:
            return

        # Get new value
        new_value = self.edit_entry.get()

        try:
            # Update tree display - with error handling
            values = list(self.tree.item(self.edit_item)['values'])
            if self.edit_col_index < len(values):
                values[self.edit_col_index] = new_value
            else:
                # Extend values list if needed
                while len(values) <= self.edit_col_index:
                    values.append("")
                values[self.edit_col_index] = new_value

            self.tree.item(self.edit_item, values=values)
        except Exception as e:
            print(f"Error updating tree item: {e}")
            # Continue with data update even if tree update fails

        # Update products_data safely - find the actual product index in original data
        # Don't use tree index as it might be different due to filtering/sorting
        if hasattr(self, 'current_product_index') and self.current_product_index >= 0:
            # Use the selected product index
            self.update_product_data_safely(self.current_product_index, self.edit_col_index, new_value)
        else:
            # Fallback: try to find by tree index (less reliable)
            try:
                item_index = self.tree.index(self.edit_item)
                if item_index < len(self.products_data):
                    self.update_product_data_safely(item_index, self.edit_col_index, new_value)
            except Exception as e:
                print(f"Error finding item index: {e}")

        # Clean up
        self.edit_entry.destroy()
        self.edit_entry = None
        self.edit_item = None
        self.edit_col_index = -1

        # Mark data as modified
        self.mark_data_modified()

        # Update default values for specific fields
        if hasattr(self, 'edit_col_index') and self.edit_col_index >= 0 and self.edit_col_index < len(self.columns):
            col_name = self.columns[self.edit_col_index]
            if col_name in self.default_values and new_value:
                self.default_values[col_name] = new_value

    def cancel_edit(self):
        """Cancel editing"""
        if self.edit_entry:
            self.edit_entry.destroy()
            self.edit_entry = None
        self.edit_item = None
        self.edit_col_index = None

    def show_column_settings(self):
        """Show enhanced column visibility settings dialog with tabs"""
        dialog = tk.Toplevel(self.root)
        dialog.title("إعدادات الأعمدة")
        dialog.geometry("1000x700")  # Doubled width like edit dialog
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg=self.purple_color)

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 500  # Adjusted for new width
        y = (dialog.winfo_screenheight() // 2) - 350
        dialog.geometry(f"+{x}+{y}")

        # Title
        title_label = tk.Label(dialog, text="إعدادات عرض الأعمدة",
                               font=self.header_font, bg=self.purple_color, fg='white')
        title_label.pack(pady=10)

        # Create notebook for tabs
        notebook = ttk.Notebook(dialog)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Basic columns tab
        basic_frame = tk.Frame(notebook, bg='white')
        notebook.add(basic_frame, text="الأعمدة الأساسية")

        # Advanced columns tab
        advanced_frame = tk.Frame(notebook, bg='white')
        notebook.add(advanced_frame, text="الأعمدة المتقدمة")

        # Basic columns (most commonly used)
        basic_columns = [
            'أسم المنتج', 'تصنيف المنتج', 'صورة المنتج', 'وصف صورة المنتج',
            'سعر المنتج', 'الكمية المتوفرة', 'رمز المنتج sku'
        ]

        # Advanced columns (less commonly used)
        advanced_columns = [col for col in self.columns if col not in basic_columns]

        # Store checkboxes
        self.column_vars = {}

        # Create basic columns checkboxes
        basic_scroll = tk.Frame(basic_frame)
        basic_scroll.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        for i, col in enumerate(basic_columns):
            var = tk.BooleanVar(value=self.visible_columns[col])
            self.column_vars[col] = var

            cb = tk.Checkbutton(basic_scroll, text=col, variable=var,
                                font=self.table_font, bg='white', anchor='w')
            cb.grid(row=i // 2, column=i % 2, sticky='w', padx=10, pady=5)

        # Create advanced columns checkboxes
        advanced_scroll = tk.Frame(advanced_frame)
        advanced_scroll.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        for i, col in enumerate(advanced_columns):
            var = tk.BooleanVar(value=self.visible_columns[col])
            self.column_vars[col] = var

            cb = tk.Checkbutton(advanced_scroll, text=col, variable=var,
                                font=self.table_font, bg='white', anchor='w')
            cb.grid(row=i // 2, column=i % 2, sticky='w', padx=10, pady=5)

        # Buttons frame
        btn_frame = tk.Frame(dialog, bg=self.purple_color)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        def apply_changes():
            # Update visibility settings
            for col, var in self.column_vars.items():
                self.visible_columns[col] = var.get()

            # Save preferences
            self.save_column_preferences()

            # Refresh table
            self.refresh_table_columns()

            # Close dialog
            dialog.destroy()

            self.show_toast("تم التطبيق", "تم تحديث إعدادات الأعمدة")

        def select_all():
            for var in self.column_vars.values():
                var.set(True)

        def select_none():
            for var in self.column_vars.values():
                var.set(False)

        def reset_defaults():
            # Reset to default visibility (all visible)
            for var in self.column_vars.values():
                var.set(True)

        # Buttons
        tk.Button(btn_frame, text="✓ تطبيق", command=apply_changes,
                  bg=self.gold_accent, fg='white', font=self.button_font).pack(side=tk.RIGHT, padx=5)
        tk.Button(btn_frame, text="❌ إلغاء", command=dialog.destroy,
                  bg='#666666', fg='white', font=self.button_font).pack(side=tk.RIGHT)

        tk.Button(btn_frame, text="🔄 افتراضي", command=reset_defaults,
                  bg=self.light_purple, fg='white', font=self.button_font).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="☑️ تحديد الكل", command=select_all,
                  bg=self.light_purple, fg='white', font=self.button_font).pack(side=tk.LEFT)
        tk.Button(btn_frame, text="☐ إلغاء الكل", command=select_none,
                  bg=self.light_purple, fg='white', font=self.button_font).pack(side=tk.LEFT)

    def refresh_table_columns(self):
        """Refresh table to show/hide columns based on visibility settings"""
        # Save current data
        current_data = []
        for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            current_data.append(values)

        # Clear tree
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Update tree columns
        visible_cols = [col for col in self.columns if self.visible_columns[col]]
        self.tree.configure(columns=visible_cols)

        # Reconfigure column headers and widths
        column_widths = {
            'أسم المنتج': 250,
            'تصنيف المنتج': 180,
            'صورة المنتج': 150,
            'وصف صورة المنتج': 200,
            'سعر المنتج': 100,
            'الكمية المتوفرة': 100,
            'رمز المنتج sku': 120,
            'الوزن': 80,
            'وحدة الوزن': 80,
            'الباركود': 120,
            'المصنعية للجرام': 100,
            'المصنعية للقطعة': 100,
            'الكمية المباعة': 100,
            'العيار': 80,
            'وزن الفصوص': 100,
            'المقاس': 80,
            'المورد': 120,
            'رقم الفاتورة': 100,
            'تاريخ الفاتورة': 100
        }

        for col in visible_cols:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_table(c))
            width = column_widths.get(col, 100)
            self.tree.column(col, width=width, minwidth=50, anchor='center', stretch=False)

        # Re-setup auto-resize functionality
        self.setup_column_auto_resize()

        # Restore data with visible columns only
        for row_data in current_data:
            visible_data = []
            for i, col in enumerate(self.columns):
                if self.visible_columns[col] and i < len(row_data):
                    visible_data.append(row_data[i])

            if visible_data:  # Only add if there's visible data
                item = self.tree.insert('', 'end', values=visible_data)
                # Apply alternating row colors
                if len(self.tree.get_children()) % 2 == 0:
                    self.tree.item(item, tags=('evenrow',))
                else:
                    self.tree.item(item, tags=('oddrow',))

        # Update row highlighting after refreshing
        self.update_row_highlighting()

    def sort_table(self, col):
        """Sort table by column"""
        data = [(self.tree.set(child, col), child) for child in self.tree.get_children('')]

        # Try to sort numerically if possible
        try:
            data.sort(key=lambda x: float(x[0]) if x[0] else 0, reverse=self.sort_reverse)
        except:
            data.sort(reverse=self.sort_reverse)

        for index, (val, child) in enumerate(data):
            self.tree.move(child, '', index)

        self.sort_reverse = not self.sort_reverse

    def show_context_menu(self, event):
        """Show context menu for table rows"""
        # Select row under cursor
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)

            # Create context menu
            menu = tk.Menu(self.root, tearoff=0)
            menu.add_command(label="✏️ تعديل", command=lambda: self.edit_product(None))
            menu.add_command(label="📋 نسخ", command=self.copy_product)
            menu.add_command(label="📋 لصق", command=self.paste_product)
            menu.add_separator()
            menu.add_command(label="🗑️ حذف", command=self.delete_product)

            menu.tk_popup(event.x_root, event.y_root)

    def copy_product(self):
        """Copy selected product data"""
        selection = self.tree.selection()
        if selection:
            item = selection[0]
            index = self.tree.index(item)
            if index < len(self.products_data):
                self.clipboard_product = self.products_data[index][:]
                if index in self.all_selected_images:
                    self.clipboard_images = self.all_selected_images[index][:]
                else:
                    self.clipboard_images = []
                self.show_toast("تم النسخ", "تم نسخ بيانات المنتج")

    def paste_product(self):
        """Paste copied product data"""
        if hasattr(self, 'clipboard_product'):
            self.save_state_for_undo()

            # Add new product with copied data
            new_index = len(self.products_data)
            self.products_data.append(self.clipboard_product[:])

            # Copy images if available
            if hasattr(self, 'clipboard_images') and self.clipboard_images:
                self.all_selected_images[new_index] = self.clipboard_images[:]

            # Refresh table
            self.refresh_product_table()
            self.mark_data_modified()
            self.show_toast("تم اللصق", "تم لصق بيانات المنتج")

    def delete_product(self):
        """Delete selected product"""
        selection = self.tree.selection()
        if selection:
            if messagebox.askyesno("تأكيد الحذف", "هل تريد حذف المنتج المحدد؟"):
                self.save_state_for_undo()

                item = selection[0]
                index = self.tree.index(item)

                # Remove from data
                if index < len(self.products_data):
                    self.products_data.pop(index)

                # Remove images
                if index in self.all_selected_images:
                    del self.all_selected_images[index]

                # Update indices for remaining images
                new_images = {}
                for k, v in self.all_selected_images.items():
                    if k > index:
                        new_images[k - 1] = v
                    elif k < index:
                        new_images[k] = v
                self.all_selected_images = new_images

                # Refresh table
                self.refresh_product_table()
                self.mark_data_modified()

    def refresh_product_table(self):
        """Refresh the product table display"""
        print(f"refresh_product_table called - products_data length: {len(self.products_data)}")  # Debug

        # Clear table
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Re-add all products with visible columns only
        for i, product in enumerate(self.products_data):
            # Get visible data only
            visible_data = []
            for j, col in enumerate(self.columns):
                if self.visible_columns[col] and j < len(product):
                    # FIXED: Special handling for barcode display (column index 9 = "الباركود")
                    if j == 9 and product[j]:  # Barcode column
                        # Ensure barcode is displayed with leading zeros preserved
                        barcode_value = str(product[j]).strip()
                        # Remove apostrophe if it exists (from Excel formatting)
                        if barcode_value.startswith("'"):
                            barcode_value = barcode_value[1:]
                        visible_data.append(barcode_value)
                    else:
                        visible_data.append(product[j])
                elif self.visible_columns[col]:
                    visible_data.append("")  # Add empty string for missing data

            # Determine tag for styling
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'

            # Check if this product has images for highlighting
            if i in self.all_selected_images and len(self.all_selected_images[i]) > 0:
                tag = 'has_images'

            self.tree.insert('', 'end', values=visible_data, tags=(tag,))

        print(f"refresh_product_table completed - added {len(self.products_data)} products")  # Debug

        # Update row highlighting to show products with images
        self.update_row_highlighting()
        self.update_status_bar()

    def _bind_mousewheel(self, widget):
        """Bind mousewheel events for scrolling"""

        def _on_mousewheel(event):
            widget.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _on_mousewheel_linux(event):
            if event.num == 4:
                widget.yview_scroll(-1, "units")
            elif event.num == 5:
                widget.yview_scroll(1, "units")

        if platform.system() == "Windows":
            widget.bind("<MouseWheel>", _on_mousewheel)
        else:
            widget.bind("<Button-4>", _on_mousewheel_linux)
            widget.bind("<Button-5>", _on_mousewheel_linux)

    def setup_layout(self):
        """Setup the initial layout"""
        self.add_empty_row()
        # Restore pane positions after a short delay to ensure UI is ready
        self.root.after(500, self.restore_pane_positions)

    def restore_pane_positions(self):
        """Restore saved pane positions"""
        if not hasattr(self, 'saved_pane_positions'):
            return

        try:
            # Restore main paned window position
            if hasattr(self, 'main_paned') and 'main_paned' in self.saved_pane_positions:
                self.main_paned.sash_place(0, self.saved_pane_positions['main_paned'], 0)

            # Restore right paned window position
            if hasattr(self, 'right_paned') and 'right_paned' in self.saved_pane_positions:
                self.right_paned.sash_place(0, self.saved_pane_positions['right_paned'], 0)

        except Exception as e:
            logger.error(f"Error restoring pane positions: {e}")

    def load_default_values(self):
        """Load default values from file if exists"""
        try:
            config_file = 'jewelry_config.json'
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.default_values = config.get('defaults', self.default_values)
                    # Load other preferences
                    if 'window_geometry' in config:
                        self.root.geometry(config['window_geometry'])
                    # Store pane positions for later restoration
                    self.saved_pane_positions = config.get('pane_positions', {})
        except Exception as e:
            logger.error(f"Error loading configuration: {e}")
            self.saved_pane_positions = {}

    def save_column_preferences(self):
        """Save column visibility preferences to file"""
        try:
            preferences = {
                'visible_columns': self.visible_columns
            }
            with open('column_preferences.json', 'w', encoding='utf-8') as f:
                json.dump(preferences, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving column preferences: {e}")

    def load_column_preferences(self):
        """Load column visibility preferences from file"""
        try:
            if os.path.exists('column_preferences.json'):
                with open('column_preferences.json', 'r', encoding='utf-8') as f:
                    preferences = json.load(f)
                    if 'visible_columns' in preferences:
                        # Update only existing columns to handle new columns in updates
                        for col in self.columns:
                            if col in preferences['visible_columns']:
                                self.visible_columns[col] = preferences['visible_columns'][col]
        except Exception as e:
            print(f"Error loading column preferences: {e}")
            # Keep default values if loading fails

    def save_default_values(self):
        """Save default values and preferences to file"""
        try:
            # Get current pane positions
            pane_positions = {}
            if hasattr(self, 'main_paned'):
                try:
                    pane_positions['main_paned'] = self.main_paned.sash_coord(0)[0]
                except:
                    pane_positions['main_paned'] = 350  # Default
            if hasattr(self, 'right_paned'):
                try:
                    pane_positions['right_paned'] = self.right_paned.sash_coord(0)[0]
                except:
                    pane_positions['right_paned'] = 300  # Default

            config = {
                'defaults': self.default_values,
                'window_geometry': self.root.geometry(),
                'last_folder': self.current_folder,
                'pane_positions': pane_positions
            }
            with open('jewelry_config.json', 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Error saving configuration: {e}")

    def select_folder(self):
        """Select folder containing images"""
        folder = filedialog.askdirectory(
            title="اختر مجلد الصور",
            initialdir=self.current_folder or os.getcwd()
        )
        if folder:
            self.current_folder = folder
            self.folder_label.config(text=f"📁 {os.path.basename(folder)}")

            # Load images in background thread
            progress = ProgressDialog(self.root, "جاري التحميل", "جاري تحميل الصور...")

            def load_task():
                self.load_images_from_folder()
                self.root.after(0, progress.close)
                self.root.after(0, self.load_existing_data)

            thread = threading.Thread(target=load_task)
            thread.daemon = True
            thread.start()

    def load_images_from_folder(self):
        """Load and display images from the selected folder with caching"""
        if not self.current_folder:
            return

        # Clear existing images
        self.root.after(0, self._clear_image_widgets)

        # Supported image formats
        image_extensions = ['*.jpg', '*.jpeg', '*.png', '*.gif', '*.bmp', '*.tiff', '*.webp']
        self.image_files = []

        for ext in image_extensions:
            pattern = os.path.join(self.current_folder, ext)
            self.image_files.extend(glob.glob(pattern))
            pattern_upper = os.path.join(self.current_folder, ext.upper())
            self.image_files.extend(glob.glob(pattern_upper))

        # Remove duplicates and sort
        self.image_files = sorted(list(set(self.image_files)))

        # Load images in batches to avoid UI freeze
        self._load_images_batch(0)

    def _clear_image_widgets(self):
        """Clear existing image widgets"""
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.image_widgets = []

    def _load_images_batch(self, start_index, batch_size=20):
        """Load images in batches for better performance"""
        if start_index >= len(self.image_files):
            self.update_all_image_borders()
            self.update_status_bar()
            return

        end_index = min(start_index + batch_size, len(self.image_files))

        row = start_index // 3
        col = start_index % 3

        for i in range(start_index, end_index):
            image_path = self.image_files[i]
            try:
                # Check cache first
                photo = self.image_cache.get(image_path, (80, 80))

                if not photo:
                    # Create thumbnail
                    img = Image.open(image_path)
                    img.thumbnail((80, 80), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                    self.image_cache.set(image_path, (80, 80), photo)

                # Create frame for image
                img_frame = tk.Frame(self.scrollable_frame, relief=tk.RAISED, borderwidth=2, bg='white')
                img_frame.grid(row=row, column=col, padx=5, pady=5)

                # Create label with image
                img_label = tk.Label(img_frame, image=photo, cursor="hand2", bg='white')
                img_label.image = photo
                img_label.pack()

                # Truncate filename for display
                filename = os.path.basename(image_path)
                display_name = filename[:15] + "..." if len(filename) > 15 else filename

                # Image name label with tooltip
                name_label = tk.Label(img_frame, text=display_name, font=('Arial', 8), bg='white')
                name_label.pack()

                # Create tooltip for full filename
                self._create_tooltip(name_label, filename)

                # Bind events
                img_label.bind('<Button-1>', lambda e, path=image_path: self.select_image(path))
                img_label.bind('<Double-Button-1>', lambda e, path=image_path: self.show_full_image(path))
                img_label.bind('<Button-3>', lambda e, path=image_path: self.show_image_context_menu(e, path))

                # Store references
                img_label.frame = img_frame
                img_label.image_path = image_path
                self.image_widgets.append(img_label)

                col += 1
                if col >= 3:
                    col = 0
                    row += 1

            except Exception as e:
                logger.error(f"Error loading image {image_path}: {e}")

        # Schedule next batch
        self.root.after(10, lambda: self._load_images_batch(end_index))

    def _create_tooltip(self, widget, text):
        """Create tooltip for widget"""

        def on_enter(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root + 10}+{event.y_root + 10}")
            label = tk.Label(tooltip, text=text, background="lightyellow",
                             relief="solid", borderwidth=1, font=('Arial', 9))
            label.pack()
            widget.tooltip = tooltip

        def on_leave(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                del widget.tooltip

        widget.bind("<Enter>", on_enter)
        widget.bind("<Leave>", on_leave)

    def show_image_context_menu(self, event, image_path):
        """Show context menu for image"""
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="📋 نسخ مسار الصورة",
                         command=lambda: self.copy_to_clipboard(image_path))
        menu.add_command(label="🔍 عرض الصورة كاملة",
                         command=lambda: self.show_full_image(image_path))
        menu.add_separator()
        menu.add_command(label="➕ إضافة للمنتج الحالي",
                         command=lambda: self.select_image(image_path))

        menu.tk_popup(event.x_root, event.y_root)

    def load_existing_data(self):
        """Load existing Excel data if available"""
        if not self.current_folder:
            return

        # Look for ANY existing Excel files (not just Products*.xlsx)
        excel_files = glob.glob(os.path.join(self.current_folder, "*.xlsx"))
        excel_files.extend(glob.glob(os.path.join(self.current_folder, "*.xls")))

        # Filter out temporary files
        excel_files = [f for f in excel_files if not os.path.basename(f).startswith('~')]

        if excel_files:
            # If there are multiple files, prefer Products*.xlsx, otherwise use any Excel file
            products_files = [f for f in excel_files if os.path.basename(f).startswith('Products')]
            if products_files:
                latest_file = max(products_files, key=os.path.getmtime)
            else:
                # Use any Excel file found
                latest_file = max(excel_files, key=os.path.getmtime)

            print(f"Loading existing Excel file: {os.path.basename(latest_file)}")  # Debug

            try:
                # COMPLETELY REWRITTEN: Load data using openpyxl only to preserve exact formatting
                print(f"Loading existing Excel file: {os.path.basename(latest_file)}")

                # Load workbook using openpyxl to preserve exact formatting
                wb = openpyxl.load_workbook(latest_file, data_only=True)
                ws = wb.active

                # Read headers from first row
                self.original_headers = []
                for cell in ws[1]:
                    self.original_headers.append(cell.value if cell.value is not None else "")

                print(f"Original headers: {self.original_headers}")

                # Create mapping between program columns and original file positions
                self.original_column_mapping = {}
                for i, program_col in enumerate(self.columns):
                    # Find this program column in the original headers
                    for j, orig_header in enumerate(self.original_headers[1:], 1):  # Skip 'No.' column
                        if str(orig_header).strip() == program_col.strip():
                            self.original_column_mapping[i] = j
                            print(f"Mapped program column {i} ({program_col}) to file column {j} ({orig_header})")
                            break
                    else:
                        # If not found, map to the expected position (fallback)
                        self.original_column_mapping[i] = i + 1
                        print(f"Fallback: Mapped program column {i} ({program_col}) to file column {i + 1}")

                # Clear existing data
                self.products_data = []
                self.all_selected_images = {}

                # Process each row starting from row 2 (skip header)
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=0):
                    if not row or not row[0]:  # Skip empty rows
                        continue

                    # Extract product data using the column mapping
                    product_row = [""] * len(self.columns)

                    for prog_col_idx, file_col_idx in self.original_column_mapping.items():
                        if file_col_idx < len(row) and row[file_col_idx] is not None:
                            value = row[file_col_idx]

                            # FIXED: Special handling for barcode column (program index 9 = "الباركود")
                            if prog_col_idx == 9:
                                # Preserve barcode exactly as stored in Excel
                                barcode_str = str(value).strip()
                                # Remove apostrophe if it was added for formatting
                                if barcode_str.startswith("'"):
                                    barcode_str = barcode_str[1:]
                                product_row[prog_col_idx] = barcode_str
                            else:
                                product_row[prog_col_idx] = str(value)

                    self.products_data.append(product_row)

                    # Process selected images if available (column index 2 = "صورة المنتج")
                    if len(product_row) > 2 and product_row[2]:
                        image_names = [name.strip() for name in str(product_row[2]).split(';')]
                        image_paths = []

                        for img_name in image_names:
                            if img_name:
                                # Find full path
                                for img_path in self.image_files:
                                    if os.path.basename(img_path) == img_name:
                                        image_paths.append(img_path)
                                        break

                        if image_paths:
                            self.all_selected_images[len(self.products_data) - 1] = image_paths

                # Close workbook
                wb.close()

                # Store the loaded file path for future saves
                self.current_excel_file = latest_file

                # Refresh display
                self.refresh_product_table()
                self.show_toast("تم التحميل ✅",
                                f"تم تحميل {len(self.products_data)} منتج من {os.path.basename(latest_file)}")

            except Exception as e:
                print(f"Error loading Excel file: {e}")
                messagebox.showerror("خطأ", f"فشل في تحميل ملف الإكسل: {e}")
        else:
            print("No Excel files found, will create new file when saving")  # Debug
            self.current_excel_file = None
            # Initialize empty mapping for new files
            self.original_column_mapping = {i: i + 1 for i in range(len(self.columns))}
            self.original_headers = ['No.'] + self.columns

    def select_image(self, image_path):
        """Select an image for the current product"""
        if self.current_product_index < 0:
            messagebox.showwarning("تحذير", "يرجى اختيار منتج أولاً")
            return

        self.save_state_for_undo()

        # Add image to current product
        if self.current_product_index not in self.all_selected_images:
            self.all_selected_images[self.current_product_index] = []

        # Allow multiple products to use the same image
        if image_path not in self.all_selected_images[self.current_product_index]:
            self.all_selected_images[self.current_product_index].append(image_path)
            self.mark_data_modified()

        # Update image borders and display
        self.update_all_image_borders()
        self.display_selected_images()
        self.update_status_bar()

        # Update row highlighting to show products with images
        self.update_row_highlighting()

    def select_all_images(self):
        """Select all images for current product"""
        if self.current_product_index < 0:
            messagebox.showwarning("تحذير", "يرجى اختيار منتج أولاً")
            return

        self.save_state_for_undo()

        self.all_selected_images[self.current_product_index] = self.image_files[:]
        self.mark_data_modified()

        self.update_all_image_borders()
        self.display_selected_images()
        self.update_status_bar()

        # Update row highlighting
        self.update_row_highlighting()

    def clear_selection(self):
        """Clear all image selections for current product"""
        if self.current_product_index < 0:
            return

        if self.current_product_index in self.all_selected_images:
            self.save_state_for_undo()
            self.all_selected_images[self.current_product_index] = []
            self.mark_data_modified()

            self.update_all_image_borders()
            self.display_selected_images()
            self.update_status_bar()

            # Update row highlighting
            self.update_row_highlighting()

    def show_full_image(self, image_path):
        """Show full size image in a popup window"""
        try:
            popup = tk.Toplevel(self.root)
            popup.title(f"عرض: {os.path.basename(image_path)}")

            # Load image
            img = Image.open(image_path)

            # Calculate appropriate size
            screen_width = popup.winfo_screenwidth()
            screen_height = popup.winfo_screenheight()
            max_width = int(screen_width * 0.8)
            max_height = int(screen_height * 0.8)

            # Resize if necessary
            img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)

            # Set window size
            popup.geometry(f"{img.width + 20}x{img.height + 40}")

            # Center window
            popup.update_idletasks()
            x = (screen_width // 2) - (popup.winfo_width() // 2)
            y = (screen_height // 2) - (popup.winfo_height() // 2)
            popup.geometry(f"+{x}+{y}")

            # Display image
            label = tk.Label(popup, image=photo)
            label.image = photo
            label.pack(expand=True, fill=tk.BOTH)

            # Bind Escape to close
            popup.bind('<Escape>', lambda e: popup.destroy())

        except Exception as e:
            messagebox.showerror("خطأ", f"لا يمكن عرض الصورة:\n{str(e)}")

    def update_all_image_borders(self):
        """Update image borders to show selected status"""
        for widget in self.image_widgets:
            if hasattr(widget, 'frame') and hasattr(widget, 'image_path'):
                # Check if image is selected for any product
                is_selected = False
                for product_images in self.all_selected_images.values():
                    if widget.image_path in product_images:
                        is_selected = True
                        break

                if is_selected:
                    widget.frame.configure(relief=tk.SOLID, borderwidth=3, bg='#4CAF50')
                else:
                    widget.frame.configure(relief=tk.RAISED, borderwidth=2, bg='white')

    def display_selected_images(self):
        """Display selected images for current product"""
        # Clear existing display
        for widget in self.selected_images_frame.winfo_children():
            widget.destroy()

        # Update info label
        if self.current_product_index >= 0:
            product_name = ""
            if self.current_product_index < len(self.products_data):
                product_name = self.products_data[self.current_product_index][
                                   0] or f"منتج {self.current_product_index + 1}"
            self.selected_info_label.config(text=f"المنتج: {product_name}")
        else:
            self.selected_info_label.config(text="لم يتم اختيار منتج")

        if self.current_product_index not in self.all_selected_images:
            no_images_label = tk.Label(self.selected_images_frame,
                                       text="لا توجد صور محددة\n\nانقر على صورة من القسم الأيسر لإضافتها",
                                       font=('Arial', 10), foreground='gray', bg='white', justify=tk.CENTER)
            no_images_label.pack(pady=20)
            return

        selected_paths = self.all_selected_images[self.current_product_index]

        if not selected_paths:
            no_images_label = tk.Label(self.selected_images_frame,
                                       text="لا توجد صور محددة\n\nانقر على صورة من القسم الأيسر لإضافتها",
                                       font=('Arial', 10), foreground='gray', bg='white', justify=tk.CENTER)
            no_images_label.pack(pady=20)
            return

        for i, image_path in enumerate(selected_paths):
            try:
                # Check cache
                photo = self.image_cache.get(image_path, (100, 100))

                if not photo:
                    img = Image.open(image_path)
                    img.thumbnail((100, 100), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                    self.image_cache.set(image_path, (100, 100), photo)

                # Create frame
                img_frame = tk.Frame(self.selected_images_frame, relief=tk.RAISED, borderwidth=1, bg='white')
                img_frame.pack(pady=5, fill=tk.X)

                # Image label
                img_label = tk.Label(img_frame, image=photo, cursor="hand2", bg='white')
                img_label.image = photo
                img_label.pack(side=tk.LEFT, padx=5)

                # Image info
                info_frame = tk.Frame(img_frame, bg='white')
                info_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

                tk.Label(info_frame, text=f"#{i + 1}", font=('Arial', 10, 'bold'), bg='white').pack(anchor='w')
                tk.Label(info_frame, text=os.path.basename(image_path),
                         font=('Arial', 8), wraplength=180, bg='white').pack(anchor='w')

                # Bind events
                img_label.bind('<Button-1>', lambda e, idx=i: self.select_image_for_removal(idx))
                img_frame.bind('<Button-1>', lambda e, idx=i: self.select_image_for_removal(idx))
                img_label.bind('<Button-3>', lambda e, path=image_path: self.show_selected_image_menu(e, path))

                img_frame.image_index = i

            except Exception as e:
                logger.error(f"Error displaying selected image: {e}")

    def show_selected_image_menu(self, event, image_path):
        """Show context menu for selected image"""
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="📋 نسخ المسار",
                         command=lambda: self.copy_to_clipboard(image_path))
        menu.add_command(label="🔍 عرض كاملة",
                         command=lambda: self.show_full_image(image_path))
        menu.add_separator()
        menu.add_command(label="❌ إزالة من المنتج",
                         command=self.remove_selected_image)

        menu.tk_popup(event.x_root, event.y_root)

    def select_image_for_removal(self, index):
        """Select an image for removal"""
        self.selected_image_for_removal = index

        # Update visual selection
        for widget in self.selected_images_frame.winfo_children():
            if hasattr(widget, 'image_index'):
                if widget.image_index == index:
                    widget.configure(bg='#E3F2FD', relief=tk.SOLID, borderwidth=2)
                else:
                    widget.configure(bg='white', relief=tk.RAISED, borderwidth=1)

    def remove_selected_image(self):
        """Remove selected image from current product"""
        if (self.current_product_index not in self.all_selected_images or
                self.selected_image_for_removal < 0):
            return

        self.save_state_for_undo()

        if self.selected_image_for_removal < len(self.all_selected_images[self.current_product_index]):
            removed = self.all_selected_images[self.current_product_index].pop(self.selected_image_for_removal)
            self.selected_image_for_removal = -1
            self.mark_data_modified()

            self.display_selected_images()
            self.update_all_image_borders()
            self.update_status_bar()

            # Update row highlighting
            self.update_row_highlighting()

            self.show_toast("تم الحذف", f"تم إزالة {os.path.basename(removed)}")

    def delete_selected_image(self):
        """Delete selected image using Delete key"""
        self.remove_selected_image()

    def reorder_images(self):
        """Show dialog to reorder images"""
        if self.current_product_index not in self.all_selected_images:
            messagebox.showinfo("تنبيه", "لا توجد صور لترتيبها")
            return

        images = self.all_selected_images[self.current_product_index]
        if len(images) < 2:
            messagebox.showinfo("تنبيه", "يجب وجود صورتين على الأقل للترتيب")
            return

        # Create reorder dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("ترتيب الصور")
        dialog.geometry("400x500")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg=self.purple_color)

        # Listbox for images
        listbox = tk.Listbox(dialog, selectmode=tk.SINGLE, height=20, font=self.table_font)
        listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        for i, img_path in enumerate(images):
            listbox.insert(tk.END, f"{i + 1}. {os.path.basename(img_path)}")

        # Buttons frame
        btn_frame = tk.Frame(dialog, bg=self.purple_color)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)

        def move_up():
            selection = listbox.curselection()
            if selection and selection[0] > 0:
                idx = selection[0]
                images[idx], images[idx - 1] = images[idx - 1], images[idx]

                # Update listbox
                listbox.delete(0, tk.END)
                for i, img_path in enumerate(images):
                    listbox.insert(tk.END, f"{i + 1}. {os.path.basename(img_path)}")
                listbox.selection_set(idx - 1)

        def move_down():
            selection = listbox.curselection()
            if selection and selection[0] < len(images) - 1:
                idx = selection[0]
                images[idx], images[idx + 1] = images[idx + 1], images[idx]

                # Update listbox
                listbox.delete(0, tk.END)
                for i, img_path in enumerate(images):
                    listbox.insert(tk.END, f"{i + 1}. {os.path.basename(img_path)}")
                listbox.selection_set(idx + 1)

        def save_order():
            self.save_state_for_undo()
            self.all_selected_images[self.current_product_index] = images
            self.mark_data_modified()
            self.display_selected_images()
            dialog.destroy()

        tk.Button(btn_frame, text="↑ أعلى", command=move_up,
                  bg=self.light_purple, fg='white', font=self.button_font).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="↓ أسفل", command=move_down,
                  bg=self.light_purple, fg='white', font=self.button_font).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="✓ حفظ", command=save_order,
                  bg=self.gold_accent, fg='white', font=self.button_font).pack(side=tk.RIGHT, padx=5)
        tk.Button(btn_frame, text="✗ إلغاء", command=dialog.destroy,
                  bg='#666666', fg='white', font=self.button_font).pack(side=tk.RIGHT)

    def on_tree_change(self, event=None):
        """Handle changes in the tree view"""
        # DON'T sync tree data - this causes data corruption
        # self.sync_tree_to_products_data()  # REMOVED - this was causing data corruption
        self.mark_data_modified()

    def update_row_highlighting(self):
        """Update row highlighting based on products with selected images - FIXED VERSION"""
        try:
            for item in self.tree.get_children():
                # Get the row index in the tree
                row_index = self.tree.index(item)

                # Check if this product (by index) has selected images
                has_images = row_index in self.all_selected_images and len(self.all_selected_images[row_index]) > 0

                if has_images:
                    # Apply pink highlighting for products with images
                    self.tree.item(item, tags=('has_images',))
                else:
                    # Apply normal alternating row colors
                    if row_index % 2 == 0:
                        self.tree.item(item, tags=('evenrow',))
                    else:
                        self.tree.item(item, tags=('oddrow',))
        except Exception as e:
            print(f"Error updating row highlighting: {e}")

    def on_product_select(self, event):
        """Handle product selection in the table"""
        selection = self.tree.selection()
        if selection:
            item = selection[0]
            # FIXED: Get the actual product index from original data, not tree display index
            tree_index = self.tree.index(item)

            # If we're in search mode, map tree index to original data index
            if hasattr(self, 'is_searching') and self.is_searching and hasattr(self, 'filtered_indices'):
                if tree_index < len(self.filtered_indices):
                    self.current_product_index = self.filtered_indices[tree_index]
                else:
                    self.current_product_index = -1
            else:
                # Normal mode - tree index matches data index
                self.current_product_index = tree_index

            self.selected_image_for_removal = -1
            print(f"Selected product: tree_index={tree_index}, actual_index={self.current_product_index}")  # Debug

            self.display_selected_images()

    def on_search_changed(self, *args):
        """Handle search text change - COMPLETELY FIXED VERSION"""
        search_text = self.search_var.get().lower().strip()
        print(f"Search changed: '{search_text}' - Original data length: {len(self.products_data)}")  # Debug message

        # Clear current display
        for item in self.tree.get_children():
            self.tree.delete(item)

        if not search_text:
            # Show all products when search is empty - use refresh_product_table
            print("Search is empty, showing all products via refresh_product_table")  # Debug message
            self.is_searching = False  # Not searching anymore
            self.filtered_indices = []  # Clear filtered indices
            self.refresh_product_table()
            return

        # We are now in search mode
        self.is_searching = True
        self.filtered_indices = []  # Track original indices of filtered products

        # Filter products based on search - but DON'T modify original data
        print(f"Filtering products with search: '{search_text}'")  # Debug message
        matches_found = 0

        for i, product in enumerate(self.products_data):
            # Search in all visible fields
            visible_data = []
            for j, col in enumerate(self.columns):
                if self.visible_columns[col] and j < len(product):
                    visible_data.append(product[j])

            # Check if search term matches any field
            if any(search_text in str(field).lower() for field in visible_data):
                # Track the original index of this product
                self.filtered_indices.append(i)

                # Add this product to display
                tag = 'evenrow' if matches_found % 2 == 0 else 'oddrow'

                # Check if this product has images for highlighting
                if i in self.all_selected_images and len(self.all_selected_images[i]) > 0:
                    tag = 'has_images'

                self.tree.insert('', 'end', values=visible_data, tags=(tag,))
                matches_found += 1

        print(f"Found {matches_found} matches - Original data still: {len(self.products_data)}")  # Debug message
        self.update_status_bar()

    def add_new_product(self):
        """Add a new empty product row"""
        self.save_state_for_undo()

        # Create new product with default values
        new_product = [""] * len(self.columns)

        # Set default values
        for i, col in enumerate(self.columns):
            if col in self.default_values and self.default_values[col]:
                new_product[i] = self.default_values[col]
            elif col == 'وحدة الوزن':
                new_product[i] = 'gm'

        self.products_data.append(new_product)

        # Refresh table
        self.refresh_product_table()

        # Select the new product
        items = self.tree.get_children()
        if items:
            last_item = items[-1]
            self.tree.selection_set(last_item)
            self.tree.focus(last_item)
            self.tree.see(last_item)

        self.mark_data_modified()
        self.show_toast("تم الإضافة", "تم إضافة منتج جديد")

    def add_empty_row(self):
        """Add an empty row for new product entry"""
        if not self.products_data or any(field.strip() for field in self.products_data[-1]):
            self.add_new_product()

    def edit_product(self, event):
        """Edit product data in an enhanced dialog"""
        selection = self.tree.selection()
        if not selection:
            return

        item = selection[0]
        product_index = self.tree.index(item)
        current_values = self.tree.item(item)['values']

        self.create_edit_dialog(product_index, current_values)

    def create_edit_dialog(self, product_index, current_values):
        """Create enhanced edit dialog with better Arabic text support"""
        dialog = tk.Toplevel(self.root)
        dialog.title("تعديل بيانات المنتج")
        dialog.geometry("1200x700")  # Doubled width from 600 to 1200 (2x)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg=self.purple_color)

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 600  # Adjusted for new width (1200/2 = 600)
        y = (dialog.winfo_screenheight() // 2) - 350
        dialog.geometry(f"+{x}+{y}")

        # Create notebook for tabbed interface
        notebook = ttk.Notebook(dialog)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Basic info tab
        basic_frame = tk.Frame(notebook, bg='white')
        notebook.add(basic_frame, text="البيانات الأساسية")

        # Pricing tab
        pricing_frame = tk.Frame(notebook, bg='white')
        notebook.add(pricing_frame, text="التسعير والكميات")

        # Technical tab
        technical_frame = tk.Frame(notebook, bg='white')
        notebook.add(technical_frame, text="البيانات التقنية")

        # Store entry widgets
        entries = {}

        # Basic info fields
        basic_fields = [
            ('أسم المنتج', 0),
            ('تصنيف المنتج', 1),
            ('وصف صورة المنتج', 3),
            ('رمز المنتج sku', 6),
            ('المقاس', 15),
            ('المورد', 16)
        ]

        # Create basic info widgets
        for i, (field_name, field_index) in enumerate(basic_fields):
            row = i // 2
            col = (i % 2) * 2

            tk.Label(basic_frame, text=field_name + ":", font=self.table_font, bg='white').grid(
                row=row, column=col, sticky='e', padx=(10, 5), pady=10)

            if field_name == 'تصنيف المنتج':
                entry = ttk.Combobox(basic_frame, values=self.product_categories,
                                     font=self.table_font, width=25)
                if field_index < len(current_values):
                    entry.set(current_values[field_index])
            else:
                entry = tk.Entry(basic_frame, font=self.table_font, width=30)
                if field_index < len(current_values):
                    entry.insert(0, str(current_values[field_index]))

            entry.grid(row=row, column=col + 1, sticky='w', padx=(5, 10), pady=10)
            entries[field_index] = entry

        # Pricing fields
        pricing_fields = [
            ('سعر المنتج', 4),
            ('الكمية المتوفرة', 5),
            ('الكمية المباعة', 12),
            ('المصنعية للجرام', 10),
            ('المصنعية للقطعة', 11)
        ]

        # Create pricing widgets
        for i, (field_name, field_index) in enumerate(pricing_fields):
            row = i // 2
            col = (i % 2) * 2

            tk.Label(pricing_frame, text=field_name + ":", font=self.table_font, bg='white').grid(
                row=row, column=col, sticky='e', padx=(10, 5), pady=10)

            entry = tk.Entry(pricing_frame, font=self.table_font, width=30)
            if field_index < len(current_values):
                entry.insert(0, str(current_values[field_index]))

            entry.grid(row=row, column=col + 1, sticky='w', padx=(5, 10), pady=10)
            entries[field_index] = entry

        # Technical fields
        technical_fields = [
            ('الوزن', 7),
            ('وحدة الوزن', 8),
            ('الباركود', 9),
            ('العيار', 13),
            ('وزن الفصوص', 14),
            ('رقم الفاتورة', 17),
            ('تاريخ الفاتورة', 18)
        ]

        # Create technical widgets
        for i, (field_name, field_index) in enumerate(technical_fields):
            row = i // 2
            col = (i % 2) * 2

            tk.Label(technical_frame, text=field_name + ":", font=self.table_font, bg='white').grid(
                row=row, column=col, sticky='e', padx=(10, 5), pady=10)

            if field_name == 'وحدة الوزن':
                entry = tk.Entry(technical_frame, font=self.table_font, width=30)
                entry.insert(0, 'gm')
                entry.configure(state='readonly')
            else:
                entry = tk.Entry(technical_frame, font=self.table_font, width=30)
                if field_index < len(current_values):
                    entry.insert(0, str(current_values[field_index]))

            entry.grid(row=row, column=col + 1, sticky='w', padx=(5, 10), pady=10)
            entries[field_index] = entry

        # Buttons frame
        btn_frame = tk.Frame(dialog, bg=self.purple_color)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        def save_changes():
            self.save_state_for_undo()

            # Update product data
            for field_index, entry in entries.items():
                if field_index < len(self.products_data[product_index]):
                    self.products_data[product_index][field_index] = entry.get()
                else:
                    # Extend list if needed
                    while len(self.products_data[product_index]) <= field_index:
                        self.products_data[product_index].append("")
                    self.products_data[product_index][field_index] = entry.get()

            # Update default values
            for field_index, entry in entries.items():
                field_name = self.columns[field_index]
                if field_name in self.default_values and entry.get():
                    self.default_values[field_name] = entry.get()

            # Refresh table
            self.refresh_product_table()
            self.mark_data_modified()

            dialog.destroy()
            self.show_toast("تم الحفظ", "تم حفظ تعديلات المنتج")

        def apply_defaults():
            # Apply default values to empty fields
            for field_index, entry in entries.items():
                field_name = self.columns[field_index]
                if field_name in self.default_values and not entry.get():
                    entry.delete(0, tk.END)
                    entry.insert(0, self.default_values[field_name])

        # Buttons
        tk.Button(btn_frame, text="✓ حفظ", command=save_changes,
                  bg=self.gold_accent, fg='white', font=self.button_font).pack(side=tk.RIGHT, padx=5)
        tk.Button(btn_frame, text="❌ إلغاء", command=dialog.destroy,
                  bg='#666666', fg='white', font=self.button_font).pack(side=tk.RIGHT)
        tk.Button(btn_frame, text="🔄 تطبيق الافتراضي", command=apply_defaults,
                  bg=self.light_purple, fg='white', font=self.button_font).pack(side=tk.LEFT, padx=5)

        # Focus on first field
        if entries:
            first_entry = entries[min(entries.keys())]
            first_entry.focus()
            first_entry.select_range(0, tk.END)

    def focus_search(self):
        """Focus on search box"""
        for widget in self.root.winfo_children():
            if isinstance(widget, tk.Entry) and widget.cget('textvariable') == str(self.search_var):
                widget.focus()
                widget.select_range(0, tk.END)
                break

    def export_data(self):
        """Export data with options"""
        if not self.products_data:
            messagebox.showwarning("تحذير", "لا توجد بيانات للتصدير")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("خيارات التصدير")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg=self.purple_color)

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 200
        y = (dialog.winfo_screenheight() // 2) - 150
        dialog.geometry(f"+{x}+{y}")

        # Title
        tk.Label(dialog, text="خيارات التصدير", font=self.header_font,
                 bg=self.purple_color, fg='white').pack(pady=10)

        # Options frame
        options_frame = tk.Frame(dialog, bg='white')
        options_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Export options
        include_images_var = tk.BooleanVar(value=True)
        include_empty_var = tk.BooleanVar(value=False)

        tk.Checkbutton(options_frame, text="تضمين أسماء الصور المحددة",
                       variable=include_images_var, font=self.table_font, bg='white').pack(anchor='w', pady=5)
        tk.Checkbutton(options_frame, text="تضمين المنتجات الفارغة",
                       variable=include_empty_var, font=self.table_font, bg='white').pack(anchor='w', pady=5)

        # Format selection
        tk.Label(options_frame, text="تنسيق التصدير:", font=self.table_font, bg='white').pack(anchor='w', pady=(10, 5))

        format_var = tk.StringVar(value="excel")
        tk.Radiobutton(options_frame, text="Excel (.xlsx)", variable=format_var, value="excel",
                       font=self.table_font, bg='white').pack(anchor='w')
        tk.Radiobutton(options_frame, text="CSV (.csv)", variable=format_var, value="csv",
                       font=self.table_font, bg='white').pack(anchor='w')

        # Buttons
        btn_frame = tk.Frame(dialog, bg=self.purple_color)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        def do_export():
            include_images = include_images_var.get()
            include_empty = include_empty_var.get()
            export_format = format_var.get()

            if export_format == "excel":
                filename = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    initialfile=f"Products_Export_{datetime.now().strftime('%d%m%Y')}.xlsx"
                )
                if filename:
                    self.export_to_excel(filename, include_images, include_empty)
            else:
                filename = filedialog.asksaveasfilename(
                    defaultextension=".csv",
                    filetypes=[("CSV files", "*.csv")],
                    initialfile=f"Products_Export_{datetime.now().strftime('%d%m%Y')}.csv"
                )
                if filename:
                    self.export_to_csv(filename, include_images, include_empty)

            dialog.destroy()

        tk.Button(btn_frame, text="📤 تصدير", command=do_export,
                  bg=self.gold_accent, fg='white', font=self.button_font).pack(side=tk.RIGHT, padx=5)
        tk.Button(btn_frame, text="❌ إلغاء", command=dialog.destroy,
                  bg='#666666', fg='white', font=self.button_font).pack(side=tk.RIGHT)

    def validate_column_mapping(self):
        """Validate and fix column mapping to ensure data integrity"""
        try:
            # Ensure we have a valid mapping for all program columns
            for i in range(len(self.columns)):
                if i not in self.original_column_mapping:
                    # Create fallback mapping
                    self.original_column_mapping[i] = i + 1
                    print(f"Added fallback mapping: program column {i} -> file column {i + 1}")

            # Ensure no duplicate file column indices
            used_file_cols = set()
            for prog_col, file_col in self.original_column_mapping.items():
                if file_col in used_file_cols:
                    # Fix duplicate mapping
                    new_file_col = max(used_file_cols) + 1 if used_file_cols else 1
                    print(
                        f"Fixed duplicate mapping: program column {prog_col} moved from file column {file_col} to {new_file_col}")
                    self.original_column_mapping[prog_col] = new_file_col
                    file_col = new_file_col
                used_file_cols.add(file_col)

            print(f"Column mapping validated: {self.original_column_mapping}")

        except Exception as e:
            print(f"Error validating column mapping: {e}")
            # Create safe fallback mapping
            self.original_column_mapping = {i: i + 1 for i in range(len(self.columns))}

    def export_to_excel(self, filename=None, include_images=True, include_empty=False):
        """SACRED FILE APPROACH: Preserve Excel file structure absolutely - only modify cell contents"""
        try:
            # FIXED: Validate column mapping before export to ensure data integrity
            self.validate_column_mapping()

            print(f"Exporting {len(self.products_data)} products to Excel")
            print(f"Using column mapping: {self.original_column_mapping}")

            # Determine filename
            if filename is None:
                if self.current_excel_file and os.path.exists(self.current_excel_file):
                    filename = self.current_excel_file
                    print(f"Updating existing file: {os.path.basename(filename)}")
                else:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = os.path.join(self.current_folder, f"Products_{timestamp}.xlsx")
                    print(f"Creating new file: {os.path.basename(filename)}")

            # SACRED FILE APPROACH: Load existing file and preserve everything
            if self.current_excel_file and os.path.exists(self.current_excel_file):
                # Load existing workbook - PRESERVE EVERYTHING
                wb = openpyxl.load_workbook(self.current_excel_file)
                ws = wb.active

                print("SACRED FILE MODE: Preserving existing file structure completely")

                # Get the current number of data rows (excluding header)
                existing_rows = 0
                for row in ws.iter_rows(min_row=2):
                    if row[0].value is not None:  # If first column (No.) has value
                        existing_rows += 1
                    else:
                        break

                print(f"Found {existing_rows} existing rows in file")

                # Update existing rows with current data (preserve row order completely)
                exported_count = 0
                for i, product in enumerate(self.products_data):
                    if not include_empty and not any(str(field).strip() for field in product):
                        continue

                    row_num = exported_count + 2  # Start from row 2 (after header)

                    # Update row number in first column (preserve existing numbering style)
                    ws.cell(row=row_num, column=1, value=exported_count + 1)

                    # Update each cell according to original column mapping
                    for prog_col_idx, file_col_idx in self.original_column_mapping.items():
                        if prog_col_idx < len(product):
                            value = product[prog_col_idx]

                            # Special handling for image column
                            if prog_col_idx == 2 and include_images and i in self.all_selected_images:
                                image_names = [os.path.basename(p) for p in self.all_selected_images[i]]
                                value = '; '.join(image_names)

                            # Get the target cell (file_col_idx is already 1-based from mapping)
                            # FIXED: Add 1 to file_col_idx because column 1 is reserved for row numbers
                            target_cell = ws.cell(row=row_num, column=file_col_idx + 1)

                            # SACRED CONTENT UPDATE: Only change cell content, preserve format
                            if prog_col_idx == 9 and value is not None and str(value).strip():
                                # Special handling for barcode - preserve as text
                                barcode_value = str(value).strip()
                                target_cell.value = barcode_value
                                target_cell.number_format = '@'  # Force text format
                            elif value is not None and str(value).strip():
                                # Normal value update - preserve existing format when possible
                                target_cell.value = str(value)
                            else:
                                # Clear cell if no value
                                target_cell.value = ""

                    exported_count += 1

                # Clear any remaining rows that had data before but don't have data now
                if exported_count < existing_rows:
                    for row_num in range(exported_count + 2, existing_rows + 2):
                        for col in range(1, len(self.original_headers) + 1):
                            ws.cell(row=row_num, column=col, value="")

            else:
                # Create new workbook only if no existing file
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Products"

                # Add headers exactly as they should be
                headers = ['No.'] + self.columns
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD",
                                                            end_color="DDDDDD",
                                                            fill_type="solid")

                # Add data for new file
                exported_count = 0
                for i, product in enumerate(self.products_data):
                    if not include_empty and not any(str(field).strip() for field in product):
                        continue

                    row_num = exported_count + 2

                    # Set row number
                    ws.cell(row=row_num, column=1, value=exported_count + 1)

                    # Set data using simple column order (for new files)
                    for prog_col_idx, value in enumerate(product):
                        if prog_col_idx < len(self.columns):
                            # Special handling for image column
                            if prog_col_idx == 2 and include_images and i in self.all_selected_images:
                                image_names = [os.path.basename(p) for p in self.all_selected_images[i]]
                                value = '; '.join(image_names)

                            target_cell = ws.cell(row=row_num, column=prog_col_idx + 2)  # +2 because column 1 is No.

                            if prog_col_idx == 9 and value is not None and str(value).strip():
                                # Barcode handling
                                barcode_value = str(value).strip()
                                target_cell.value = barcode_value
                                target_cell.number_format = '@'
                            elif value is not None and str(value).strip():
                                target_cell.value = str(value)
                            else:
                                target_cell.value = ""

                    exported_count += 1

            # Save file
            wb.save(filename)
            wb.close()

            # Update current file reference
            self.current_excel_file = filename

            print(f"SACRED FILE: Successfully updated {exported_count} products while preserving file structure")
            return exported_count

        except Exception as e:
            print(f"Error in export_to_excel: {e}")
            raise e

    def export_to_csv(self, filename, include_images=True, include_empty=False):
        """Export data to CSV format"""
        try:
            # DON'T sync tree data - use original data directly to preserve order
            # self.sync_tree_to_products_data()  # REMOVED - this was causing data corruption

            data = []
            data.append(['No.'] + self.columns)  # Headers

            exported_count = 0
            for i, product in enumerate(self.products_data):
                if not include_empty and not any(str(field).strip() for field in product):
                    continue

                row_data = [exported_count + 1] + product.copy()

                if include_images and i in self.all_selected_images:
                    image_names = [os.path.basename(p) for p in self.all_selected_images[i]]
                    row_data[3] = '; '.join(image_names)

                data.append(row_data)
                exported_count += 1

            df = pd.DataFrame(data[1:], columns=data[0])
            df.to_csv(filename, index=False, encoding='utf-8-sig')

            self.show_toast("تم التصدير ✅", f"تم تصدير {exported_count} منتج إلى CSV")
        except Exception as e:
            messagebox.showerror("خطأ", f"فشل التصدير:\n{str(e)}")

    def save_data(self):
        """Save all product data - use existing file or create new one"""
        if not self.current_folder:
            folder = filedialog.askdirectory(title="اختر مجلد للحفظ")
            if not folder:
                return False
            self.current_folder = folder

        try:
            # Use existing file if available, otherwise create new one
            if self.current_excel_file and os.path.exists(self.current_excel_file):
                filepath = self.current_excel_file
                print(f"Saving to existing file: {os.path.basename(filepath)}")  # Debug
            else:
                # Create new file with timestamp
                timestamp = datetime.now().strftime("%d%m%Y")
                filename = f"Products {timestamp}.xlsx"
                filepath = os.path.join(self.current_folder, filename)
                print(f"Creating new file: {os.path.basename(filepath)}")  # Debug

                # Check if file exists
                if os.path.exists(filepath):
                    result = messagebox.askyesno("الملف موجود",
                                                 f"الملف {filename} موجود بالفعل.\nهل تريد استبداله؟")
                    if not result:
                        # Ask for new filename
                        filepath = filedialog.asksaveasfilename(
                            initialdir=self.current_folder,
                            defaultextension=".xlsx",
                            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                            title="احفظ باسم"
                        )
                        if not filepath:
                            return False

            # Use the updated export_to_excel function (which doesn't sync tree data)
            saved_count = self.export_to_excel(filepath, include_images=True, include_empty=False)

            self.mark_data_saved()
            self.show_toast("تم الحفظ ✅", f"تم حفظ {saved_count} منتج في {os.path.basename(filepath)}")
            return True

        except Exception as e:
            print(f"Error in save_data: {e}")
            messagebox.showerror("خطأ في الحفظ", f"حدث خطأ أثناء حفظ البيانات:\n{str(e)}")
            return False

    def update_product_data_safely(self, product_index, field_index, new_value):
        """Update product data safely without corrupting order"""
        try:
            if 0 <= product_index < len(self.products_data):
                # Ensure the product row has enough columns
                while len(self.products_data[product_index]) <= field_index:
                    self.products_data[product_index].append("")

                # FIXED: Special handling for barcode field (field_index 9 = "الباركود")
                if field_index == 9 and new_value:
                    # Preserve leading zeros in barcode
                    barcode_value = str(new_value).strip()
                    # Remove apostrophe if user accidentally entered it
                    if barcode_value.startswith("'"):
                        barcode_value = barcode_value[1:]
                    self.products_data[product_index][field_index] = barcode_value
                    print(f"Updated product {product_index}, barcode field to: {barcode_value}")  # Debug
                else:
                    # Update the specific field normally
                    self.products_data[product_index][field_index] = new_value
                    print(f"Updated product {product_index}, field {field_index} to: {new_value}")  # Debug

        except Exception as e:
            print(f"Error updating product data: {e}")

    def sync_tree_to_products_data(self):
        """Sync tree view data to products_data array"""
        # Don't sync if we're in search mode - this would corrupt the original data!
        if hasattr(self, 'is_searching') and self.is_searching:
            print("Skipping sync_tree_to_products_data because we're in search mode")  # Debug
            return

        print("Syncing tree to products_data")  # Debug
        self.products_data = []
        for item in self.tree.get_children():
            values = list(self.tree.item(item)['values'])
            # Ensure we have the right number of columns
            while len(values) < len(self.columns):
                values.append('')
            self.products_data.append(values)

    def show_advanced_search(self):
        """Show advanced search dialog like Excel"""
        dialog = tk.Toplevel(self.root)
        dialog.title("البحث المتقدم")
        dialog.geometry("800x600")
        dialog.transient(self.root)

        # Style headers
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True, size=12)
        for cell in ws[2]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD",
                                                    end_color="DDDDDD",
                                                    fill_type="solid")

        # DON'T sync tree data - use original data directly to preserve order
        # self.sync_tree_to_products_data()  # REMOVED - this was causing data corruption

        # Add data
        saved_count = 0
        for i, product in enumerate(self.products_data):
            if not include_empty and not any(str(field).strip() for field in product):
                continue

            row_data = [saved_count + 1] + product.copy()

            # Add selected images
            if include_images and i in self.all_selected_images and self.all_selected_images[i]:
                image_names = []
                for img_path in self.all_selected_images[i]:
                    image_names.append(os.path.basename(img_path))
                row_data[3] = '; '.join(image_names)  # Image column

            ws.append(row_data)
            saved_count += 1

        # Auto-adjust column widths for better Arabic text display
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 5, 60)  # Increased padding for Arabic text
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(filepath)

        # Update state
        self.mark_data_saved()
        self.save_default_values()

        if show_message:
            self.show_toast("تم الحفظ ✅", f"تم حفظ {saved_count} منتج")

    def show_advanced_search(self):
        """Show advanced search dialog like Excel"""
        dialog = tk.Toplevel(self.root)
        dialog.title("البحث المتقدم")
        dialog.geometry("800x600")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg=self.purple_color)

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 400
        y = (dialog.winfo_screenheight() // 2) - 300
        dialog.geometry(f"+{x}+{y}")

        # Title
        title_label = tk.Label(dialog, text="البحث في جميع البيانات",
                               font=self.header_font, bg=self.purple_color, fg='white')
        title_label.pack(pady=10)

        # Search frame
        search_frame = tk.Frame(dialog, bg=self.purple_color)
        search_frame.pack(fill=tk.X, padx=10, pady=5)

        tk.Label(search_frame, text="البحث عن:", font=self.table_font,
                 bg=self.purple_color, fg='white').pack(side=tk.LEFT, padx=5)

        search_entry = tk.Entry(search_frame, font=self.table_font, width=25)
        search_entry.pack(side=tk.LEFT, padx=5)

        search_btn = tk.Button(search_frame, text="🔍 بحث",
                               bg=self.gold_accent, fg='white', font=self.button_font,
                               width=8)
        search_btn.pack(side=tk.LEFT, padx=2)

        clear_btn = tk.Button(search_frame, text="🗑️ إلغاء",
                              bg='#FF6B6B', fg='white', font=self.button_font,
                              width=8)
        clear_btn.pack(side=tk.LEFT, padx=2)

        # Results frame
        results_frame = tk.Frame(dialog, bg='white')
        results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Results table
        results_columns = ['رقم الصف', 'العمود', 'القيمة', 'السياق']
        results_tree = ttk.Treeview(results_frame, columns=results_columns, show='headings', height=15)

        # Configure columns
        results_tree.heading('رقم الصف', text='رقم الصف')
        results_tree.heading('العمود', text='العمود')
        results_tree.heading('القيمة', text='القيمة')
        results_tree.heading('السياق', text='السياق')

        results_tree.column('رقم الصف', width=80, anchor='center')
        results_tree.column('العمود', width=150, anchor='center')
        results_tree.column('القيمة', width=200, anchor='center')
        results_tree.column('السياق', width=300, anchor='center')

        # Scrollbars for results
        v_scroll = ttk.Scrollbar(results_frame, orient="vertical", command=results_tree.yview)
        h_scroll = ttk.Scrollbar(results_frame, orient="horizontal", command=results_tree.xview)
        results_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        results_tree.grid(row=0, column=0, sticky='nsew')
        v_scroll.grid(row=0, column=1, sticky='ns')
        h_scroll.grid(row=1, column=0, sticky='ew')

        results_frame.grid_rowconfigure(0, weight=1)
        results_frame.grid_columnconfigure(0, weight=1)

        # Status label
        status_label = tk.Label(dialog, text="أدخل كلمة البحث واضغط بحث",
                                font=self.table_font, bg=self.purple_color, fg='white')
        status_label.pack(pady=5)

        # Create search variable for text change tracking
        search_var = tk.StringVar()
        search_entry.config(textvariable=search_var)

        def on_search_text_change(*args):
            """Auto-clear search when text is empty"""
            current_text = search_var.get().strip()
            if not current_text and self.search_results:
                # Text was cleared, auto-clear search
                clear_search_silent()

        def clear_search_silent():
            """Clear search without showing toast message"""
            # Clear results table
            for item in results_tree.get_children():
                results_tree.delete(item)

            # Clear search state
            self.search_results = []
            self.current_search_index = -1
            self.last_search_term = ""

            # Update status
            status_label.config(text="جميع المنتجات ظاهرة")

            # Clear selection in main tree
            if self.tree.selection():
                self.tree.selection_remove(self.tree.selection())

        def clear_search():
            """Clear search results and show all products"""
            # Clear search entry
            search_entry.delete(0, tk.END)

            # Clear search silently
            clear_search_silent()

            # Show feedback
            self.show_toast("تم الإلغاء", "تم إلغاء البحث وإظهار جميع المنتجات")

        # Bind text change event
        search_var.trace('w', on_search_text_change)

        def perform_search():
            search_term = search_entry.get().strip()
            if not search_term:
                messagebox.showwarning("تحذير", "يرجى إدخال كلمة البحث")
                return

            # Clear previous results
            for item in results_tree.get_children():
                results_tree.delete(item)

            results = []

            # Search in all data
            for row_index, item in enumerate(self.tree.get_children()):
                values = self.tree.item(item)['values']
                visible_cols = [col for col in self.columns if self.visible_columns[col]]

                for col_index, col_name in enumerate(visible_cols):
                    if col_index < len(values):
                        cell_value = str(values[col_index]).lower()
                        if search_term.lower() in cell_value:
                            # Get context (product name for reference)
                            context = ""
                            if len(values) > 0:
                                context = f"المنتج: {values[0]}"  # First column is usually product name

                            results.append({
                                'row': row_index + 1,
                                'column': col_name,
                                'value': values[col_index],
                                'context': context,
                                'tree_item': item
                            })

            # Display results
            for result in results:
                results_tree.insert('', 'end', values=(
                    result['row'],
                    result['column'],
                    result['value'],
                    result['context']
                ))

            # Update status
            count = len(results)
            if count > 0:
                status_label.config(text=f"تم العثور على {count} نتيجة")
                self.search_results = results
                self.current_search_index = -1
                self.last_search_term = search_term
            else:
                status_label.config(text="لم يتم العثور على نتائج")
                self.search_results = []

        def go_to_result(event):
            """Navigate to selected result in main table"""
            selection = results_tree.selection()
            if selection and self.search_results:
                item = results_tree.item(selection[0])
                row_num = int(item['values'][0]) - 1  # Convert to 0-based index

                # Find the corresponding result
                for i, result in enumerate(self.search_results):
                    if result['row'] == row_num + 1:
                        tree_item = result['tree_item']

                        # Select and focus the item in main tree
                        self.tree.selection_set(tree_item)
                        self.tree.focus(tree_item)
                        self.tree.see(tree_item)

                        # Update search index for F3/Shift+F3
                        self.current_search_index = i

                        # Close search dialog
                        dialog.destroy()

                        # Show feedback
                        self.show_toast("تم الانتقال", f"تم الانتقال إلى الصف {row_num + 1}")
                        break

        # Bind events
        search_btn.configure(command=perform_search)
        clear_btn.configure(command=clear_search)
        search_entry.bind('<Return>', lambda e: perform_search())
        results_tree.bind('<Double-1>', go_to_result)

        # Buttons frame
        button_frame = tk.Frame(dialog, bg=self.purple_color)
        button_frame.pack(fill=tk.X, padx=10, pady=10)

        tk.Button(button_frame, text="❌ إغلاق", command=dialog.destroy,
                  bg='#666666', fg='white', font=self.button_font).pack(side=tk.RIGHT, padx=5)

        # Focus on search entry
        search_entry.focus()

    def find_next(self):
        """Find next occurrence (F3)"""
        if self.search_results and self.current_search_index < len(self.search_results) - 1:
            self.current_search_index += 1
            self.navigate_to_search_result()
        else:
            self.show_toast("البحث", "لا توجد نتائج أخرى")

    def find_previous(self):
        """Find previous occurrence (Shift+F3)"""
        if self.search_results and self.current_search_index > 0:
            self.current_search_index -= 1
            self.navigate_to_search_result()
        else:
            self.show_toast("البحث", "لا توجد نتائج سابقة")

    def navigate_to_search_result(self):
        """Navigate to current search result"""
        if self.search_results and 0 <= self.current_search_index < len(self.search_results):
            result = self.search_results[self.current_search_index]
            tree_item = result['tree_item']

            # Select and focus the item
            self.tree.selection_set(tree_item)
            self.tree.focus(tree_item)
            self.tree.see(tree_item)

            # Show feedback
            current = self.current_search_index + 1
            total = len(self.search_results)
            self.show_toast("البحث", f"النتيجة {current} من {total}")

    def clear_search_filter(self):
        """Clear search filter and show all products"""
        try:
            print("Clear search button clicked!")  # Debug message

            # Clear the search entry widget directly
            if hasattr(self, 'search_entry'):
                self.search_entry.delete(0, tk.END)

            # Also clear the search variable
            self.search_var.set("")

            # Force focus away from search entry to trigger any pending events
            self.root.focus()

            # Manually trigger search change
            self.on_search_changed()

            print(f"Search cleared, products should show: {len(self.products_data)}")  # Debug message

        except Exception as e:
            print(f"Error in clear_search_filter: {e}")
            messagebox.showerror("خطأ", f"حدث خطأ في مسح البحث: {e}")

    def _create_tooltip_for_widget(self, widget, text):
        """Create tooltip for any widget"""

        def on_enter(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root + 10}+{event.y_root + 10}")
            label = tk.Label(tooltip, text=text, background="lightyellow",
                             relief="solid", borderwidth=1, font=('Arial', 9))
            label.pack()
            widget.tooltip = tooltip

        def on_leave(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                del widget.tooltip

        widget.bind("<Enter>", on_enter)
        widget.bind("<Leave>", on_leave)


def main():
    """Main function to run the application"""
    try:
        # Set DPI awareness for Windows
        if platform.system() == "Windows":
            try:
                from ctypes import windll
                windll.shcore.SetProcessDpiAwareness(1)
            except:
                pass

        root = tk.Tk()

        app = JewelryStoreManager(root)

        # Center the window
        root.update_idletasks()
        x = (root.winfo_screenwidth() // 2) - (root.winfo_width() // 2)
        y = (root.winfo_screenheight() // 2) - (root.winfo_height() // 2)
        root.geometry(f"+{x}+{y}")

        root.mainloop()

    except Exception as e:
        logger.error(f"Application error: {e}")
        messagebox.showerror("خطأ في التطبيق", f"حدث خطأ في التطبيق:\n{str(e)}")


if __name__ == "__main__":
    main()


    def clear_search_filter(self):
        """Clear search filter and show all products"""
        try:
            print("Clear search button clicked!")  # Debug message

            # Clear the search entry widget directly
            if hasattr(self, 'search_entry'):
                self.search_entry.delete(0, tk.END)

            # Also clear the search variable
            self.search_var.set("")

            # Force focus away from search entry to trigger any pending events
            self.root.focus()

            # Manually trigger search change
            self.on_search_changed()

            print(f"Search cleared, products should show: {len(self.products_data)}")  # Debug message

        except Exception as e:
            print(f"Error in clear_search_filter: {e}")
            messagebox.showerror("خطأ", f"حدث خطأ في مسح البحث: {e}")


    def _create_tooltip_for_widget(self, widget, text):
        """Create tooltip for any widget"""

        def on_enter(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root + 10}+{event.y_root + 10}")
            label = tk.Label(tooltip, text=text, background="lightyellow",
                             relief="solid", borderwidth=1, font=('Arial', 9))
            label.pack()
            widget.tooltip = tooltip

        def on_leave(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                del widget.tooltip

        widget.bind("<Enter>", on_enter)
        widget.bind("<Leave>", on_leave)


